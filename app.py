import os, random, json, io, math, uuid, re, secrets, hashlib, smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'instance', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app = Flask(__name__)
app.secret_key = "sts_secret_karabakh_2024"
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{os.path.join(BASE_DIR, 'instance', 'smart_timetable.db')}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

db = SQLAlchemy(app)

# ─────────────── MODELS ───────────────
class University(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    domain = db.Column(db.String(100), unique=True, nullable=False)
    address = db.Column(db.String(300))
    founded = db.Column(db.Integer)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(200), nullable=False)
    email = db.Column(db.String(200), unique=True, nullable=False)
    password_hash = db.Column(db.String(300), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # student / teacher / admin
    university_id = db.Column(db.Integer, db.ForeignKey("university.id"))
    department = db.Column(db.String(100))
    year = db.Column(db.Integer)
    avatar_initials = db.Column(db.String(5))
    group_id = db.Column(db.Integer, db.ForeignKey("group.id"), nullable=True)

class Course(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), nullable=False)
    name = db.Column(db.String(200), nullable=False)
    credits = db.Column(db.Integer, default=3)
    is_lab = db.Column(db.Boolean, default=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey("user.id"))
    university_id = db.Column(db.Integer, db.ForeignKey("university.id"))
    color = db.Column(db.String(20), default="#0f8f5f")

class Room(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    capacity = db.Column(db.Integer, default=30)
    is_lab = db.Column(db.Boolean, default=False)
    university_id = db.Column(db.Integer, db.ForeignKey("university.id"))

class Group(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    year = db.Column(db.Integer, default=1)
    size = db.Column(db.Integer, default=25)
    university_id = db.Column(db.Integer, db.ForeignKey("university.id"))

class Enrollment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("user.id"))
    course_id = db.Column(db.Integer, db.ForeignKey("course.id"))
    grade = db.Column(db.Float)
    attendance = db.Column(db.Float, default=100.0)
    status = db.Column(db.String(20), default="active")

class ScheduleEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey("course.id"))
    teacher_id = db.Column(db.Integer, db.ForeignKey("user.id"))
    room_id = db.Column(db.Integer, db.ForeignKey("room.id"))
    group_id = db.Column(db.Integer, db.ForeignKey("group.id"))
    day = db.Column(db.String(20))
    time_start = db.Column(db.String(10))
    time_end = db.Column(db.String(10))
    semester = db.Column(db.String(20))
    generated_at = db.Column(db.DateTime, default=datetime.utcnow)

class GeneratedTimetable(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    day = db.Column(db.String(20))
    start_time = db.Column(db.String(10))
    end_time = db.Column(db.String(10))
    course_id = db.Column(db.String(20))
    course_name = db.Column(db.String(200))
    teacher_name = db.Column(db.String(200))
    room_name = db.Column(db.String(50))
    group_name = db.Column(db.String(50))
    course_type = db.Column(db.String(20))
    status = db.Column(db.String(20), default='scheduled')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class UploadedTimetableFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(300))
    uploaded_by = db.Column(db.Integer, db.ForeignKey("user.id"))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    analysis_status = db.Column(db.String(20), default='pending')

class SystemSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    is_timetable_published = db.Column(db.Boolean, default=False)

class CourseMaterial(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey("course.id"), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.String(500))
    file_type = db.Column(db.String(20), default="pdf")  # pdf, ppt, doc, link, video
    file_url = db.Column(db.String(500))  # URL or file path
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class ExtraLessonRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    teacher_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    course_id = db.Column(db.Integer, db.ForeignKey("course.id"), nullable=False)
    room_id = db.Column(db.Integer, db.ForeignKey("room.id"), nullable=False)
    lesson_date = db.Column(db.Date, nullable=False)
    day = db.Column(db.String(20), nullable=False)
    start_time = db.Column(db.String(10), nullable=False)
    end_time = db.Column(db.String(10), nullable=False)
    message = db.Column(db.String(500))
    status = db.Column(db.String(20), default="notified")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Announcement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(500), nullable=False)
    announcement_type = db.Column(db.String(20), default="info")
    audience = db.Column(db.String(20), default="all")
    source_type = db.Column(db.String(50))
    source_id = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class PasswordResetToken(db.Model):
    """Stores SHA-256 hashes of single-use password-reset tokens."""
    __tablename__ = "password_reset_token"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    token_hash = db.Column(db.String(255), nullable=False, unique=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    expires_at = db.Column(db.DateTime, nullable=False)
    used_at = db.Column(db.DateTime, nullable=True)

def get_publish_status():
    settings = SystemSettings.query.first()
    if not settings:
        settings = SystemSettings(is_timetable_published=False)
        db.session.add(settings)
        db.session.commit()
    return settings.is_timetable_published

def get_form_redirect_target(default_endpoint):
    next_url = (request.form.get("next") or "").strip()
    if next_url.startswith("/") and not next_url.startswith("//"):
        return next_url
    return url_for(default_endpoint)

def clear_generated_timetable():
    GeneratedTimetable.query.delete()
    settings = get_or_create_settings()
    settings.is_timetable_published = False

def migrate_db():
    """Safely add new columns to existing SQLite DB without breaking data."""
    from sqlalchemy import text
    with db.engine.connect() as conn:
        try:
            conn.execute(text("ALTER TABLE user ADD COLUMN group_id INTEGER REFERENCES \"group\"(id)"))
            conn.commit()
        except Exception:
            pass  # Column already exists

        # ── password_reset_token table ──────────────────────────────────────
        try:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS password_reset_token (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL REFERENCES user(id),
                    token_hash VARCHAR(255) NOT NULL UNIQUE,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    expires_at DATETIME NOT NULL,
                    used_at DATETIME
                )
            """))
            conn.commit()
        except Exception:
            pass
        # Create CourseMaterial table if not exists
        try:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS course_material (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    course_id INTEGER NOT NULL REFERENCES course(id),
                    teacher_id INTEGER NOT NULL REFERENCES user(id),
                    title VARCHAR(200) NOT NULL,
                    description VARCHAR(500),
                    file_type VARCHAR(20) DEFAULT 'pdf',
                    file_url VARCHAR(500),
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """))
            conn.commit()
        except Exception:
            pass

        try:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS extra_lesson_request (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    teacher_id INTEGER NOT NULL REFERENCES user(id),
                    course_id INTEGER NOT NULL REFERENCES course(id),
                    room_id INTEGER NOT NULL REFERENCES room(id),
                    lesson_date DATE NOT NULL,
                    day VARCHAR(20) NOT NULL,
                    start_time VARCHAR(10) NOT NULL,
                    end_time VARCHAR(10) NOT NULL,
                    message VARCHAR(500),
                    status VARCHAR(20) DEFAULT 'notified',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """))
            conn.commit()
        except Exception:
            pass
        try:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS announcement (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    title VARCHAR(500) NOT NULL,
                    announcement_type VARCHAR(20) DEFAULT 'info',
                    audience VARCHAR(20) DEFAULT 'all',
                    source_type VARCHAR(50),
                    source_id INTEGER,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            """))
            conn.commit()
        except Exception:
            pass

# SEED
def seed_database():
    if University.query.first():
        return
    uni = University(name="EMM", domain="karabakh.edu.az",
                     address="Baku, Azerbaijan", founded=2021)
    db.session.add(uni)
    db.session.flush()

    users_data = [
        ("Məhəmməd Məmmədzadə", "student@karabakh.edu.az", "student", "Computer Science", 2, "MM"),
        ("Nigar Alishzada", "teacher@karabakh.edu.az", "teacher", "Mathematics", None, "NA"),
        ("University Admin", "admin@karabakh.edu.az", "admin", "Administration", None, "UA"),
        ("Leyla Hasanova", "leyla@karabakh.edu.az", "student", "Computer Science", 2, "LH"),
        ("Rauf Mammadov", "rauf@karabakh.edu.az", "teacher", "Computer Science", None, "RM"),
        ("Anar Guliyev", "anar@karabakh.edu.az", "teacher", "Physics", None, "AG"),
    ]
    user_objs = []
    for full_name, email, role, dept, year, initials in users_data:
        u = User(full_name=full_name, email=email,
                 password_hash=generate_password_hash("password123"),
                 role=role, university_id=uni.id, department=dept,
                 year=year, avatar_initials=initials)
        db.session.add(u)
        user_objs.append(u)
    db.session.flush()

    teacher1 = User.query.filter_by(email="teacher@karabakh.edu.az").first()
    teacher2 = User.query.filter_by(email="rauf@karabakh.edu.az").first()
    teacher3 = User.query.filter_by(email="anar@karabakh.edu.az").first()

    courses_data = [
        ("MATH201", "Discrete Mathematics", 4, False, teacher1.id, "#0f8f5f"),
        ("MATH301", "Probability Theory and Statistics", 3, False, teacher1.id, "#064e3b"),
        ("CS204", "Computer Graphics", 3, True, teacher2.id, "#059669"),
        ("CS310", "Data Analysis", 3, True, teacher2.id, "#10b981"),
        ("CS320", "Web Engineering", 3, False, teacher2.id, "#34d399"),
        ("CS401", "Artificial Intelligence", 4, False, teacher3.id, "#065f46"),
        ("CS350", "Database Systems", 3, False, teacher1.id, "#047857"),
    ]
    course_objs = []
    for code, name, credits, is_lab, tid, color in courses_data:
        c = Course(code=code, name=name, credits=credits, is_lab=is_lab,
                   teacher_id=tid, university_id=uni.id, color=color)
        db.session.add(c)
        course_objs.append(c)
    db.session.flush()

    rooms_data = [
        ("A-101", 40, False), ("A-204", 35, False), ("B-118", 30, False),
        ("C-301", 50, False), ("Lab-1", 25, True), ("Lab-2", 25, True), ("Lab-3", 20, True),
    ]
    room_objs = []
    for name, cap, is_lab in rooms_data:
        r = Room(name=name, capacity=cap, is_lab=is_lab, university_id=uni.id)
        db.session.add(r)
        room_objs.append(r)
    db.session.flush()

    groups_data = [("KOM24A", 2, 28), ("KOM24B", 2, 26), ("KOM25A", 1, 30), ("KOM25B", 1, 27)]
    group_objs = []
    for name, year, size in groups_data:
        g = Group(name=name, year=year, size=size, university_id=uni.id)
        db.session.add(g)
        group_objs.append(g)
    db.session.flush()

    student = User.query.filter_by(email="student@karabakh.edu.az").first()
    student2 = User.query.filter_by(email="leyla@karabakh.edu.az").first()
    # Assign students to groups
    student.group_id = group_objs[0].id   # KOM24A
    student2.group_id = group_objs[1].id  # KOM24B
    grades = [85.0, 90.0, 78.0, 92.0, 88.0]
    for i, c in enumerate(course_objs[:5]):
        e = Enrollment(student_id=student.id, course_id=c.id,
                       grade=grades[i], attendance=round(85 + random.uniform(0, 14), 1))
        db.session.add(e)

    db.session.commit()

# ─────────────── PASSWORD-RESET HELPERS ───────────────

def _hash_reset_token(raw_token: str) -> str:
    """Return the SHA-256 hex-digest of a raw token string."""
    return hashlib.sha256(raw_token.encode()).hexdigest()


def _generate_reset_token(user) -> str:
    """
    Create a secure single-use token for *user*, persist its hash to the DB,
    and return the raw (un-hashed) token to be embedded in the reset URL.
    """
    from datetime import timedelta
    # Invalidate any outstanding unused tokens for this user
    PasswordResetToken.query.filter_by(user_id=user.id, used_at=None).delete()
    raw_token = secrets.token_urlsafe(32)
    token_row = PasswordResetToken(
        user_id=user.id,
        token_hash=_hash_reset_token(raw_token),
        expires_at=datetime.utcnow() + timedelta(hours=1),
    )
    db.session.add(token_row)
    db.session.commit()
    return raw_token


def _get_valid_reset_token(raw_token: str):
    """
    Look up the hashed token; return the PasswordResetToken row if it is
    valid (exists, not expired, not used), else return None.
    """
    token_hash = _hash_reset_token(raw_token)
    row = PasswordResetToken.query.filter_by(token_hash=token_hash).first()
    if row is None:
        return None
    if row.used_at is not None:
        return None
    if datetime.utcnow() > row.expires_at:
        return None
    return row


def _send_password_reset_email(user, reset_url: str):
    """
    Send a password-reset email via SMTP.
    If SMTP is not configured, print the reset link to the console instead
    of crashing so developers can still test locally.
    """
    mail_server = os.environ.get("MAIL_SERVER", "")
    mail_port = int(os.environ.get("MAIL_PORT", 587))
    mail_use_tls = os.environ.get("MAIL_USE_TLS", "true").lower() == "true"
    mail_username = os.environ.get("MAIL_USERNAME", "")
    mail_password = os.environ.get("MAIL_PASSWORD", "")
    mail_sender = os.environ.get("MAIL_DEFAULT_SENDER", mail_username)

    # --- Build the message -------------------------------------------------
    name = user.full_name or "User"
    subject = "Password Reset Request – Smart Timetable System"
    html_body = f"""
    <html><body style="font-family:Arial,sans-serif;color:#222;">
      <div style="max-width:560px;margin:0 auto;padding:32px 24px;">
        <h2 style="color:#0f8f5f;">Smart Timetable System</h2>
        <p>Hello <strong>{name}</strong>,</p>
        <p>We received a request to reset your password.
           Click the button below to set a new password:</p>
        <p style="margin:28px 0;">
          <a href="{reset_url}"
             style="background:#0f8f5f;color:#fff;padding:12px 28px;
                    border-radius:6px;text-decoration:none;font-weight:600;">
            Reset Password
          </a>
        </p>
        <p style="font-size:13px;color:#666;">
          Or copy this link into your browser:<br>
          <a href="{reset_url}" style="color:#0f8f5f;">{reset_url}</a>
        </p>
        <p style="font-size:13px;color:#666;">
          This link will expire in <strong>1 hour</strong>.<br>
          If you did not request a password reset, you can safely ignore this email.
        </p>
        <hr style="border:none;border-top:1px solid #eee;margin-top:32px;">
        <p style="font-size:11px;color:#aaa;">Smart Timetable System – EMM</p>
      </div>
    </body></html>
    """
    plain_body = (
        f"Hello {name},\n\n"
        "We received a request to reset your password.\n"
        "Open the link below to set a new password:\n\n"
        f"{reset_url}\n\n"
        "This link will expire in 1 hour.\n"
        "If you did not request this, you can ignore this email."
    )

    # --- Send or fall back to console print --------------------------------
    if not mail_server or not mail_username or not mail_password:
        print("\n" + "=" * 60)
        print("[DEV] SMTP not configured – password reset link:")
        print(reset_url)
        print("=" * 60 + "\n")
        return

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = mail_sender
        msg["To"] = user.email
        msg.attach(MIMEText(plain_body, "plain"))
        msg.attach(MIMEText(html_body, "html"))

        with smtplib.SMTP(mail_server, mail_port, timeout=15) as smtp:
            if mail_use_tls:
                smtp.starttls()
            smtp.login(mail_username, mail_password)
            smtp.sendmail(mail_sender, user.email, msg.as_string())
    except Exception as exc:
        print(f"[WARN] Failed to send reset email: {exc}")
        print(f"[DEV] Reset link: {reset_url}")


# ─────────────── AUTH DECORATORS ───────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if "user_id" not in session:
                return redirect(url_for("login"))
            if session.get("role") not in roles:
                flash("Access denied.", "error")
                return redirect(url_for("login"))
            return f(*args, **kwargs)
        return decorated
    return decorator

# ─────────────── ROUTES ───────────────
@app.route("/")
def index():
    if "user_id" in session:
        role = session.get("role")
        if role == "student": return redirect(url_for("student_dashboard"))
        if role == "teacher": return redirect(url_for("teacher_dashboard"))
        if role == "admin": return redirect(url_for("admin_dashboard"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "").strip()
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password_hash, password):
            uni = University.query.get(user.university_id)
            session["user_id"] = user.id
            session["role"] = user.role
            session["full_name"] = user.full_name
            session["university_name"] = uni.name if uni else "University"
            session["avatar"] = user.avatar_initials
            if user.role == "student": return redirect(url_for("student_dashboard"))
            if user.role == "teacher": return redirect(url_for("teacher_dashboard"))
            if user.role == "admin": return redirect(url_for("admin_dashboard"))
        flash("Invalid email or password.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ── FORGOT / RESET PASSWORD ──────────────────────────────────────────────────

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    """Step 1: user enters their email to request a reset link."""
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        user = User.query.filter_by(email=email).first()
        if user:
            raw_token = _generate_reset_token(user)
            base_url = os.environ.get("BASE_URL", "").rstrip("/")
            if not base_url:
                base_url = request.host_url.rstrip("/")
            reset_url = f"{base_url}/reset-password/{raw_token}"
            _send_password_reset_email(user, reset_url)
        # Always show the same message – never reveal whether the email exists
        flash(
            "If an account with that email exists, a password reset link has been sent.",
            "success",
        )
        return redirect(url_for("login"))
    return render_template("forgot_password.html")


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    """Step 2: user opens the link and sets a new password."""
    token_row = _get_valid_reset_token(token)
    if token_row is None:
        flash(
            "This password reset link is invalid, has expired, or has already been used. "
            "Please request a new one.",
            "error",
        )
        return redirect(url_for("forgot_password"))

    if request.method == "POST":
        # Re-validate in the POST path too (prevents replay after tab sits open)
        token_row = _get_valid_reset_token(token)
        if token_row is None:
            flash(
                "This link is no longer valid. Please request a new reset link.",
                "error",
            )
            return redirect(url_for("forgot_password"))

        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if new_password != confirm_password:
            flash("Passwords do not match.", "error")
            return render_template("reset_password.html", token=token)

        if len(new_password) < 8 or not re.search(r"[A-Z]", new_password):
            flash(
                "Password must be at least 8 characters and include at least 1 uppercase letter.",
                "error",
            )
            return render_template("reset_password.html", token=token)

        user = User.query.get(token_row.user_id)
        if user is None:
            flash("User not found. Please contact support.", "error")
            return redirect(url_for("login"))

        # Update password and mark token as used
        user.password_hash = generate_password_hash(new_password)
        token_row.used_at = datetime.utcnow()
        db.session.commit()

        flash("Your password has been updated. Please log in with your new password.", "success")
        return redirect(url_for("login"))

    return render_template("reset_password.html", token=token)

# ── STUDENT ──
@app.route("/student/dashboard")
@role_required("student")
def student_dashboard():
    user = User.query.get(session["user_id"])
    enrollments = Enrollment.query.filter_by(student_id=user.id).all()
    courses, grades = [], []
    avg_grade, avg_att = 0, 0
    for e in enrollments:
        c = Course.query.get(e.course_id)
        t = User.query.get(c.teacher_id) if c else None
        courses.append({"course": c, "teacher": t, "enrollment": e})
        if e.grade: grades.append(e.grade)
        if e.attendance: avg_att += e.attendance
    avg_grade = round(sum(grades) / len(grades), 1) if grades else 0
    avg_att = round(avg_att / len(enrollments), 1) if enrollments else 0
    
    is_published = get_publish_status()
    if is_published:
        if user.group_id:
            schedule = ScheduleEntry.query.filter_by(group_id=user.group_id).all()
        else:
            # fallback: filter by enrolled courses
            student_course_ids = [e.course_id for e in enrollments]
            schedule = ScheduleEntry.query.filter(ScheduleEntry.course_id.in_(student_course_ids)).all() if student_course_ids else []
    else:
        schedule = []

    entries = build_timetable_context(schedule)
    announcements = build_recent_announcements("student") + [
        {"title": "Midterm Exam Schedule Released", "date": "May 15, 2026", "type": "important"},
        {"title": "Lab sessions moved to Lab-2 this week", "date": "May 12, 2026", "type": "info"},
        {"title": "Registration deadline: May 20", "date": "May 11, 2026", "type": "warning"},
    ]
    user_group = Group.query.get(user.group_id) if user.group_id else None
    return render_template("student_dashboard.html", user=user, courses=courses,
                           avg_grade=avg_grade, avg_att=avg_att,
                           timetable=entries, announcements=announcements,
                           is_published=is_published, user_group=user_group)

@app.route("/student/timetable")
@role_required("student")
def student_timetable():
    user = User.query.get(session["user_id"])
    is_published = get_publish_status()
    user_group = Group.query.get(user.group_id) if user.group_id else None
    if is_published:
        if user.group_id:
            schedule = ScheduleEntry.query.filter_by(group_id=user.group_id).order_by(ScheduleEntry.day, ScheduleEntry.time_start).all()
        else:
            enrollments = Enrollment.query.filter_by(student_id=user.id).all()
            ids = [e.course_id for e in enrollments]
            schedule = ScheduleEntry.query.filter(ScheduleEntry.course_id.in_(ids)).order_by(ScheduleEntry.day, ScheduleEntry.time_start).all() if ids else []
    else:
        schedule = []
    entries = build_timetable_context(schedule)
    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    timetable_by_day = {day: [] for day in days_order}
    for item in entries:
        day = item["entry"].day
        if day in timetable_by_day:
            timetable_by_day[day].append(item)
    return render_template("student_timetable.html", user=user,
                           timetable_by_day=timetable_by_day, days_order=days_order,
                           is_published=is_published, total=len(entries), user_group=user_group)

# ── TEACHER ──
@app.route("/teacher/dashboard")
@role_required("teacher")
def teacher_dashboard():
    user = User.query.get(session["user_id"])
    my_courses = Course.query.filter_by(teacher_id=user.id).all()
    
    is_published = get_publish_status()
    if is_published:
        schedule = ScheduleEntry.query.filter_by(teacher_id=user.id).all()
    else:
        schedule = []
        
    entries = build_timetable_context(schedule)
    student_count = 0
    course_details = []
    for c in my_courses:
        enrolled = Enrollment.query.filter_by(course_id=c.id).count()
        student_count += enrolled
        course_details.append({"course": c, "students": enrolled})
    grade_data = []
    for c in my_courses:
        enrollments = Enrollment.query.filter_by(course_id=c.id).all()
        for e in enrollments:
            s = User.query.get(e.student_id)
            if s:
                grade_data.append({"student": s, "course": c, "grade": e.grade,
                                   "attendance": e.attendance})
    weekly_hours = len(schedule) * 1.5 if is_published else 0
    pending = sum(1 for e in Enrollment.query.all() if not e.grade)
    return render_template("teacher_dashboard.html", user=user, courses=course_details,
                           timetable=entries, grade_data=grade_data,
                           student_count=student_count, weekly_hours=weekly_hours,
                           pending=pending, is_published=is_published)

@app.route("/teacher/timetable")
@role_required("teacher")
def teacher_timetable():
    user = User.query.get(session["user_id"])
    return render_template("teacher_timetable.html", **build_teacher_timetable_context(user))

@app.route("/teacher/extra-lesson", methods=["POST"])
@role_required("teacher")
def teacher_extra_lesson():
    user = User.query.get(session["user_id"])
    action = request.form.get("action", "check")
    require_room = action == "request"
    form_data, errors = parse_extra_lesson_form(user, require_room=require_room)

    if errors:
        for error in errors:
            flash(error, "error")
        return render_template("teacher_timetable.html", **build_teacher_timetable_context(user, extra_lesson_form=form_data))

    if teacher_has_slot_conflict(user, form_data["day"], form_data["lesson_date"], form_data["start_min"], form_data["end_min"]):
        flash("You already have a teaching session during that time.", "error")
        return render_template("teacher_timetable.html", **build_teacher_timetable_context(user, extra_lesson_form=form_data))

    available_rooms = find_available_extra_lesson_rooms(
        form_data["course"],
        form_data["day"],
        form_data["lesson_date"],
        form_data["start_min"],
        form_data["end_min"],
    )

    if action == "check":
        if not available_rooms:
            flash("No rooms are available for that date and time.", "error")
        return render_template("teacher_timetable.html", **build_teacher_timetable_context(
            user,
            extra_lesson_form=form_data,
            available_rooms=available_rooms,
        ))

    selected_room = Room.query.get(form_data["room_id"])
    if not selected_room or selected_room.id not in {room["id"] for room in available_rooms}:
        flash("That room is no longer available for the selected time.", "error")
        return render_template("teacher_timetable.html", **build_teacher_timetable_context(
            user,
            extra_lesson_form=form_data,
            available_rooms=available_rooms,
        ))

    request_row = ExtraLessonRequest(
        teacher_id=user.id,
        course_id=form_data["course"].id,
        room_id=selected_room.id,
        lesson_date=form_data["lesson_date"],
        day=form_data["day"],
        start_time=form_data["start_time"],
        end_time=form_data["end_time"],
        message=form_data["message"],
        status="notified",
    )
    db.session.add(request_row)
    db.session.flush()
    db.session.add(build_extra_lesson_announcement(request_row))
    db.session.commit()
    flash(f"Admin has been notified that you will teach in room {selected_room.name}.", "success")
    return redirect(url_for("teacher_timetable"))

# ── ADMIN ──
@app.route("/admin/dashboard")
@role_required("admin")
def admin_dashboard():
    user = User.query.get(session["user_id"])
    stats = {
        "students": User.query.filter_by(role="student").count(),
        "teachers": User.query.filter_by(role="teacher").count(),
        "courses": Course.query.count(),
        "rooms": Room.query.count(),
        "groups": Group.query.count(),
        "schedule_entries": ScheduleEntry.query.count(),
    }
    all_students = User.query.filter_by(role="student").all()
    all_teachers = User.query.filter_by(role="teacher").all()
    all_courses = Course.query.all()
    announcements = build_recent_announcements("admin")
    extra_lesson_rows = build_extra_lesson_request_rows(
        ExtraLessonRequest.query.order_by(ExtraLessonRequest.created_at.desc()).limit(10).all()
    )
    return render_template("admin_dashboard.html", user=user, stats=stats,
                           all_students=all_students, all_teachers=all_teachers,
                           all_courses=all_courses,
                           announcements=announcements,
                           extra_lesson_rows=extra_lesson_rows)

@app.route("/admin/timetable")
@role_required("admin")
def timetable_builder():
    user = User.query.get(session["user_id"])
    last_upload = UploadedTimetableFile.query.filter_by(uploaded_by=user.id).order_by(UploadedTimetableFile.uploaded_at.desc()).first()
    latest_generated_message = None
    last_results = []
    if last_upload:
        if last_upload.analysis_status == "generated":
            last_results = GeneratedTimetable.query.order_by(GeneratedTimetable.created_at.desc()).limit(200).all()
        elif last_upload.analysis_status in ["uploaded", "analyzed", "issues"]:
            latest_generated_message = "No timetable has been generated for the latest uploaded file yet."
    else:
        last_results = GeneratedTimetable.query.order_by(GeneratedTimetable.created_at.desc()).limit(200).all()
    return render_template("timetable_builder.html", user=user,
                           last_upload=last_upload, last_results=last_results,
                           latest_generated_message=latest_generated_message,
                           excel_available=EXCEL_AVAILABLE)

@app.route("/admin/generate-timetable", methods=["POST"])
@role_required("admin")
def generate_timetable():
    data = request.get_json() or {}
    semester = data.get("semester", "Spring 2026")
    max_per_day = int(data.get("max_per_day", 4))
    start_h = int(data.get("start_hour", 8))
    end_h = int(data.get("end_hour", 18))

    GeneratedTimetable.query.delete()
    db.session.commit()

    courses = Course.query.all()
    teachers_map = {c.id: c.teacher_id for c in courses}
    rooms = Room.query.all()
    lab_rooms = [r for r in rooms if r.is_lab]
    normal_rooms = [r for r in rooms if not r.is_lab]
    groups = Group.query.all()
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    slots = []
    h = start_h
    while h + 2 <= end_h:
        slots.append((f"{h:02d}:00", f"{h+2:02d}:00"))
        h += 2

    # slot_used[(day, slot_idx)] = {teacher_ids, room_ids, group_ids}
    slot_used = {}
    for d in days:
        for si in range(len(slots)):
            slot_used[(d, si)] = {"teachers": set(), "rooms": set(), "groups": set()}

    entries_created = []
    day_counts = {d: 0 for d in days}

    course_group_pairs = []
    for c in courses:
        for g in groups:
            course_group_pairs.append((c, g))
    random.shuffle(course_group_pairs)

    for course, group in course_group_pairs:
        assigned = False
        random.shuffle(days)
        for day in days:
            if day_counts[day] >= max_per_day:
                continue
            for si, (ts, te) in enumerate(slots):
                key = (day, si)
                tid = teachers_map.get(course.id)
                if tid in slot_used[key]["teachers"]:
                    continue
                if group.id in slot_used[key]["groups"]:
                    continue
                pool = lab_rooms if course.is_lab else normal_rooms
                if not pool:
                    pool = rooms
                available_rooms = [r for r in pool if r.id not in slot_used[key]["rooms"]]
                if not available_rooms:
                    continue
                t_obj = User.query.get(tid) if tid else None
                entry = GeneratedTimetable(
                    course_id=course.code,
                    course_name=course.name,
                    teacher_name=t_obj.full_name if t_obj else "Unassigned Teacher",
                    room_name=room.name,
                    group_name=group.name,
                    day=day,
                    start_time=ts,
                    end_time=te,
                    course_type="lab" if course.is_lab else "lecture",
                    status="scheduled"
                )
                db.session.add(entry)
                slot_used[key]["teachers"].add(tid)
                slot_used[key]["rooms"].add(room.id)
                slot_used[key]["groups"].add(group.id)
                day_counts[day] += 1
                entries_created.append(entry)
                assigned = True
                break
            if assigned:
                break

    settings = get_or_create_settings()
    settings.is_timetable_published = False
    db.session.commit()

    generated = GeneratedTimetable.query.all()
    # Conflicts calculation is mock for this mock generation endpoint
    conflicts = []
    result = []
    for e in generated:
        result.append({
            "id": e.id, "course": e.course_name, "code": e.course_id,
            "teacher": e.teacher_name, "room": e.room_name,
            "group": e.group_name, "day": e.day,
            "time_start": e.start_time, "time_end": e.end_time,
            "is_lab": e.course_type == "lab", "color": "#7c3aed" if e.course_type == "lab" else "#0f8f5f"
        })
    return jsonify({"success": True, "entries": result,
                    "count": len(result), "conflicts": len(conflicts),
                    "semester": semester})

@app.route("/api/timetable")
@login_required
def api_timetable():
    schedule = ScheduleEntry.query.all()
    result = []
    for e in schedule:
        c = Course.query.get(e.course_id)
        t = User.query.get(e.teacher_id)
        r = Room.query.get(e.room_id)
        g = Group.query.get(e.group_id)
        result.append({
            "course": c.name if c else "", "teacher": t.full_name if t else "",
            "room": r.name if r else "", "group": g.name if g else "",
            "day": e.day, "time_start": e.time_start, "time_end": e.time_end,
        })
    return jsonify(result)

@app.route("/courses")
@login_required
def courses():
    user = User.query.get(session["user_id"])
    if session["role"] == "student":
        enrollments = Enrollment.query.filter_by(student_id=user.id).all()
        is_published = get_publish_status()
        # Get course IDs that are actually scheduled (in timetable)
        if is_published:
            scheduled_course_ids = set(e.course_id for e in ScheduleEntry.query.all())
        else:
            scheduled_course_ids = set()
        course_list = []
        for e in enrollments:
            c = Course.query.get(e.course_id)
            t = User.query.get(c.teacher_id) if c else None
            # Only show courses that appear in the published timetable
            in_schedule = c.id in scheduled_course_ids if c else False
            course_list.append({"course": c, "teacher": t, "enrollment": e, "in_schedule": in_schedule})
        # Filter to only scheduled courses if published; otherwise show all enrolled
        if is_published:
            course_list = [item for item in course_list if item["in_schedule"]]
    elif session["role"] == "teacher":
        my_courses = Course.query.filter_by(teacher_id=user.id).all()
        course_list = []
        for c in my_courses:
            t = user
            enrolled_count = Enrollment.query.filter_by(course_id=c.id).count()
            course_list.append({"course": c, "teacher": t, "enrollment": None, "enrolled_count": enrolled_count})
    else:
        all_courses = Course.query.all()
        course_list = []
        for c in all_courses:
            t = User.query.get(c.teacher_id)
            course_list.append({"course": c, "teacher": t, "enrollment": None, "enrolled_count": Enrollment.query.filter_by(course_id=c.id).count()})
    return render_template("courses.html", user=user, courses=course_list)

@app.route("/course/<int:course_id>")
@login_required
def course_detail(course_id):
    user = User.query.get(session["user_id"])
    course = Course.query.get_or_404(course_id)
    teacher = User.query.get(course.teacher_id)
    # Student-specific data
    enrollment = None
    my_schedule = []
    if session["role"] == "student":
        enrollment = Enrollment.query.filter_by(student_id=user.id, course_id=course.id).first()
        # Only show sessions for the student's own group
        if user.group_id:
            entries = ScheduleEntry.query.filter_by(course_id=course.id, group_id=user.group_id).order_by(ScheduleEntry.day, ScheduleEntry.time_start).all()
        else:
            entries = ScheduleEntry.query.filter_by(course_id=course.id).order_by(ScheduleEntry.day, ScheduleEntry.time_start).all()
        for e in entries:
            room = Room.query.get(e.room_id)
            my_schedule.append({"entry": e, "room": room})
    # All enrolled students (for teacher/admin)
    enrolled_students = []
    schedule_with_details = []
    if session["role"] in ["teacher", "admin"]:
        all_entries = ScheduleEntry.query.filter_by(course_id=course.id).order_by(ScheduleEntry.day, ScheduleEntry.time_start).all()
        for e in all_entries:
            room = Room.query.get(e.room_id)
            group = Group.query.get(e.group_id)
            schedule_with_details.append({"entry": e, "room": room, "group": group})
        enrollments = Enrollment.query.filter_by(course_id=course.id).all()
        for e in enrollments:
            s = User.query.get(e.student_id)
            if s:
                enrolled_students.append({"student": s, "enrollment": e})
    enrolled_count = Enrollment.query.filter_by(course_id=course.id).count()
    # Fetch course materials
    materials = CourseMaterial.query.filter_by(course_id=course.id).order_by(CourseMaterial.created_at.desc()).all()
    user_group = Group.query.get(user.group_id) if user.group_id else None
    return render_template("course_detail.html", user=user, course=course, teacher=teacher,
                           my_schedule=my_schedule, schedule=schedule_with_details,
                           enrollment=enrollment, enrolled_students=enrolled_students,
                           enrolled_count=enrolled_count, materials=materials,
                           user_group=user_group)

@app.route("/course/<int:course_id>/material", methods=["POST"])
@role_required("teacher")
def add_course_material(course_id):
    user = User.query.get(session["user_id"])
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != user.id:
        flash("You are not authorized to add materials to this course.", "danger")
        return redirect(url_for("course_detail", course_id=course_id))
    
    title = request.form.get("title")
    description = request.form.get("description")
    file_type = request.form.get("file_type")
    file_url = request.form.get("file_url")
    
    if title:
        material = CourseMaterial(
            course_id=course.id,
            teacher_id=user.id,
            title=title,
            description=description,
            file_type=file_type,
            file_url=file_url
        )
        db.session.add(material)
        db.session.commit()
        flash("Course material added successfully.", "success")
    return redirect(url_for("course_detail", course_id=course_id))

@app.route("/course/<int:course_id>/grade", methods=["POST"])
@role_required("teacher")
def grade_student(course_id):
    user = User.query.get(session["user_id"])
    course = Course.query.get_or_404(course_id)
    if course.teacher_id != user.id:
        flash("You are not authorized to grade students in this course.", "danger")
        return redirect(url_for("course_detail", course_id=course_id))
    
    student_id = request.form.get("student_id")
    grade = request.form.get("grade")
    attendance = request.form.get("attendance")
    
    enrollment = Enrollment.query.filter_by(course_id=course.id, student_id=student_id).first()
    if enrollment:
        if grade:
            enrollment.grade = float(grade)
        if attendance:
            enrollment.attendance = float(attendance)
        db.session.commit()
        flash("Student grades updated successfully.", "success")
    return redirect(url_for("course_detail", course_id=course_id))


@app.route("/grades")
@login_required
def grades():
    user = User.query.get(session["user_id"])
    grade_data = []
    if session["role"] == "student":
        enrollments = Enrollment.query.filter_by(student_id=user.id).all()
        for e in enrollments:
            c = Course.query.get(e.course_id)
            t = User.query.get(c.teacher_id) if c else None
            grade_data.append({"course": c, "teacher": t, "enrollment": e})
    return render_template("grades.html", user=user, grade_data=grade_data)

@app.route("/profile")
@login_required
def profile():
    user = User.query.get(session["user_id"])
    uni = University.query.get(user.university_id) if user.university_id else None
    return render_template("profile.html", user=user, university=uni)

@app.route("/change-password", methods=["POST"])
@login_required
def change_password():
    user = User.query.get(session["user_id"])
    current_password = request.form.get("current_password", "")
    new_password = request.form.get("new_password", "")
    confirm_new_password = request.form.get("confirm_new_password", "")
    
    if not check_password_hash(user.password_hash, current_password):
        flash("Current password is incorrect.", "error")
        return redirect(url_for("profile"))
        
    if new_password != confirm_new_password:
        flash("New passwords do not match.", "error")
        return redirect(url_for("profile"))
        
    if len(new_password) < 8 or not re.search(r"[A-Z]", new_password):
        flash("New password must be at least 8 characters long and include at least 1 capital letter.", "error")
        return redirect(url_for("profile"))
        
    user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    flash("Password changed successfully.", "success")
    return redirect(url_for("profile"))

@app.route("/admin/students")
@role_required("admin")
def admin_students():
    user = User.query.get(session["user_id"])
    students = User.query.filter_by(role="student").order_by(User.full_name).all()
    student_rows = build_admin_student_rows(students)
    return render_template("admin_students.html", user=user, student_rows=student_rows)

@app.route("/admin/teachers")
@role_required("admin")
def admin_teachers():
    user = User.query.get(session["user_id"])
    teachers = User.query.filter_by(role="teacher").order_by(User.full_name).all()
    teacher_rows = build_admin_teacher_rows(teachers)
    return render_template("admin_teachers.html", user=user, teacher_rows=teacher_rows)

@app.route("/admin/courses")
@role_required("admin")
def admin_courses():
    user = User.query.get(session["user_id"])
    courses = Course.query.order_by(Course.code, Course.name).all()
    teachers = User.query.filter_by(role="teacher").all()
    t_map = {t.id: t for t in teachers}
    course_rows = build_admin_course_rows(courses, t_map)
    return render_template("admin_courses.html", user=user, course_rows=course_rows, t_map=t_map)

@app.route("/admin/rooms")
@role_required("admin")
def admin_rooms():
    user = User.query.get(session["user_id"])
    rooms = Room.query.order_by(Room.name).all()
    return render_template("admin_rooms.html", user=user, rooms=rooms)

@app.route("/admin/rooms/empty", methods=["POST"])
@role_required("admin")
def admin_empty_rooms():
    data = request.get_json(silent=True) or request.form
    day = safe_text(data.get("day"))
    date_text = safe_text(data.get("date"))
    start_text = safe_text(data.get("start_time"))
    end_text = safe_text(data.get("end_time"))

    if date_text:
        try:
            selected = datetime.strptime(date_text, "%Y-%m-%d")
            day = WEEKDAYS[selected.weekday()] if selected.weekday() < len(WEEKDAYS) else selected.strftime("%A")
        except ValueError:
            return jsonify({"success": False, "error": "Please enter a valid date."}), 400

    start_min = parse_time_value(start_text)
    end_min = parse_time_value(end_text)
    if not day or start_min is None or end_min is None or start_min >= end_min:
        return jsonify({"success": False, "error": "Please select a day/date and a valid time range."}), 400

    rooms = Room.query.order_by(Room.name).all()
    empty_rooms = find_empty_rooms_for_slot(rooms, day, start_min, end_min)
    return jsonify({
        "success": True,
        "day": day,
        "start_time": format_time_minutes(start_min),
        "end_time": format_time_minutes(end_min),
        "rooms": empty_rooms,
        "count": len(empty_rooms),
    })

@app.route("/admin/groups")
@role_required("admin")
def admin_groups():
    user = User.query.get(session["user_id"])
    groups = Group.query.order_by(Group.name).all()
    group_rows = build_admin_group_rows(groups)
    return render_template("admin_groups.html", user=user, group_rows=group_rows)

@app.route("/admin/generated-timetable")
@role_required("admin")
def admin_generated_timetable():
    user = User.query.get(session["user_id"])
    last_upload = UploadedTimetableFile.query.filter_by(uploaded_by=user.id).order_by(UploadedTimetableFile.uploaded_at.desc()).first()
    if last_upload and last_upload.analysis_status in ["uploaded", "analyzed", "issues"]:
        generated_schedule = []
    else:
        generated_schedule = GeneratedTimetable.query.order_by(GeneratedTimetable.day, GeneratedTimetable.start_time).all()
    if generated_schedule:
        entries = build_generated_master_rows(generated_schedule)
        source_label = "Smart timetable draft"
    else:
        schedule = ScheduleEntry.query.order_by(ScheduleEntry.day, ScheduleEntry.time_start).all()
        entries = build_schedule_master_rows(schedule)
        source_label = "Active timetable"
    conflicts = detect_master_row_conflicts(entries)
    is_published = get_publish_status()
    return render_template("generated_timetable.html", user=user, entries=entries,
                           conflict_count=len(conflicts), is_published=is_published,
                           source_label=source_label)

@app.route("/admin/send-timetable", methods=["POST"])
@role_required("admin")
def send_timetable_to_sections():
    redirect_target = get_form_redirect_target("admin_generated_timetable")
    try:
        last_upload = UploadedTimetableFile.query.filter_by(uploaded_by=session["user_id"]).order_by(UploadedTimetableFile.uploaded_at.desc()).first()
        generated_count = GeneratedTimetable.query.count()
        if last_upload and last_upload.analysis_status in ["uploaded", "analyzed", "issues"]:
            generated_count = 0
        active_count = ScheduleEntry.query.count()
        if generated_count:
            sent_count = sync_generated_timetable_to_schedule()
        elif active_count:
            settings = get_or_create_settings()
            settings.is_timetable_published = True
            db.session.commit()
            sent_count = active_count
        else:
            flash("No timetable entries are available to send.", "error")
            return redirect(redirect_target)

        if sent_count:
            flash(f"{sent_count} class sessions were sent to the teacher and student sections.", "success")
        else:
            flash("No scheduled sessions were available to send.", "error")
    except Exception:
        db.session.rollback()
        flash("Could not send the timetable. Please check the generated rows and try again.", "error")
    return redirect(redirect_target)

@app.route("/admin/publish-timetable", methods=["POST"])
@role_required("admin")
def publish_timetable():
    settings = get_or_create_settings()
    settings.is_timetable_published = not settings.is_timetable_published
    db.session.commit()
    status = "published" if settings.is_timetable_published else "unpublished"
    flash(f"Timetable has been {status} successfully.", "success")
    return redirect(url_for("admin_generated_timetable"))

@app.route("/admin/settings")
@role_required("admin")
def admin_settings():
    user = User.query.get(session["user_id"])
    uni = University.query.first()
    return render_template("admin_settings.html", user=user, university=uni)

@app.route("/admin/change-password", methods=["POST"])
@role_required("admin")
def admin_change_password():
    user = User.query.get(session["user_id"])
    current_password = request.form.get("current_password", "")
    new_password = request.form.get("new_password", "")
    confirm_new_password = request.form.get("confirm_new_password", "")

    if not check_password_hash(user.password_hash, current_password):
        flash("Current password is incorrect.", "error")
        return redirect(url_for("admin_settings"))
    if new_password != confirm_new_password:
        flash("New passwords do not match.", "error")
        return redirect(url_for("admin_settings"))
    if len(new_password) < 8 or not re.search(r"[A-Z]", new_password):
        flash("New password must be at least 8 characters long and include at least 1 capital letter.", "error")
        return redirect(url_for("admin_settings"))

    user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    flash("Admin password updated successfully.", "success")
    return redirect(url_for("admin_settings"))

# ─────────────── EXCEL TIMETABLE ROUTES ───────────────

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
REQUIRED_EXCEL_SHEETS = [
    "Teachers", "TeacherAvailability", "Courses", "Rooms",
    "Groups", "TimeSlots", "Constraints"
]
REQUIRED_EXCEL_COLUMNS = {
    "Teachers": ["teacher_id", "teacher_name", "department", "email", "max_weekly_hours"],
    "TeacherAvailability": ["teacher_id", "day", "start_time", "end_time", "availability_status"],
    "Courses": ["course_id", "course_name", "teacher_id", "group_id", "course_type", "weekly_hours", "required_room_type"],
    "Rooms": ["room_id", "room_name", "room_type", "capacity"],
    "Groups": ["group_id", "group_name", "student_count", "department", "semester"],
    "TimeSlots": ["slot_id", "day", "start_time", "end_time"],
    "Constraints": ["constraint_name", "constraint_type", "priority", "value"],
}

def safe_upload_path(filename):
    safe_name = secure_filename(filename or "")
    if not safe_name or not allowed_file(safe_name):
        return None, None
    upload_root = os.path.abspath(app.config["UPLOAD_FOLDER"])
    path = os.path.abspath(os.path.join(upload_root, safe_name))
    if os.path.commonpath([upload_root, path]) != upload_root:
        return None, None
    return safe_name, path

def clean_cell(value):
    if EXCEL_AVAILABLE and pd.isna(value):
        return ""
    return str(value).strip()

def normalize_day(value):
    text = clean_cell(value)
    for day in WEEKDAYS:
        if text.lower() == day.lower():
            return day
    return text

def parse_time_value(value):
    if EXCEL_AVAILABLE and pd.isna(value):
        return None
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        return int(round(float(value) * 24 * 60))
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return int(value.hour) * 60 + int(value.minute)
    text = clean_cell(value)
    if not text:
        return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.hour * 60 + parsed.minute
        except ValueError:
            continue
    return None

def format_time_minutes(minutes):
    return f"{minutes // 60:02d}:{minutes % 60:02d}"

def parse_float(value, default=None):
    try:
        if EXCEL_AVAILABLE and pd.isna(value):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default

def parse_bool(value):
    return clean_cell(value).lower() in {"true", "1", "yes", "y", "enabled"}

def read_excel_sheets(filepath):
    try:
        xl = pd.ExcelFile(filepath)
    except Exception as exc:
        message = str(exc)
        if "xlrd" in message.lower():
            raise ValueError("Could not read .xls file. Install xlrd support or upload the .xlsx template.") from exc
        raise ValueError("Could not read Excel file. Please upload a valid .xlsx or .xls workbook.") from exc

    found_sheets = set(xl.sheet_names)
    sheets = {}
    errors = []

    for sheet_name in REQUIRED_EXCEL_SHEETS:
        if sheet_name not in found_sheets:
            errors.append(f"Excel file is missing the {sheet_name} sheet.")
            continue
        df = xl.parse(sheet_name)
        df.columns = [clean_cell(col) for col in df.columns]
        sheets[sheet_name] = df
        for column in REQUIRED_EXCEL_COLUMNS[sheet_name]:
            if column not in df.columns:
                errors.append(f"{sheet_name} sheet is missing {column} column.")

    return sheets, errors

def sheet_has_columns(sheets, sheet_name):
    if sheet_name not in sheets:
        return False
    return all(col in sheets[sheet_name].columns for col in REQUIRED_EXCEL_COLUMNS[sheet_name])

def get_column_values(df, column):
    return [clean_cell(value) for value in df[column].tolist()]

def get_available_windows(avail_df):
    windows = {}
    if avail_df is None:
        return windows
    for _, row in avail_df.iterrows():
        status = clean_cell(row.get("availability_status")).lower()
        if status and status != "available":
            continue
        teacher_id = clean_cell(row.get("teacher_id"))
        day = normalize_day(row.get("day"))
        start = parse_time_value(row.get("start_time"))
        end = parse_time_value(row.get("end_time"))
        if teacher_id and day and start is not None and end is not None and start < end:
            windows.setdefault(teacher_id, {}).setdefault(day, []).append((start, end))
    return windows

def get_timeslots(slots_df):
    slots = []
    if slots_df is None:
        return slots
    seen = set()
    for _, row in slots_df.iterrows():
        day = normalize_day(row.get("day"))
        start = parse_time_value(row.get("start_time"))
        end = parse_time_value(row.get("end_time"))
        slot_id = clean_cell(row.get("slot_id"))
        if not day or start is None or end is None or start >= end:
            continue
        key = (day, start, end)
        if key in seen:
            continue
        seen.add(key)
        slots.append({
            "id": slot_id,
            "day": day,
            "start_min": start,
            "end_min": end,
            "start": format_time_minutes(start),
            "end": format_time_minutes(end),
            "duration_hours": (end - start) / 60,
        })
    day_order = {day: idx for idx, day in enumerate(WEEKDAYS)}
    slots.sort(key=lambda item: (day_order.get(item["day"], 99), item["start_min"], item["end_min"]))
    return slots

def validate_excel_data(sheets, schema_errors=None):
    errors = list(schema_errors or [])
    warnings = []

    summary = {
        "teachers": len(sheets["Teachers"]) if "Teachers" in sheets else 0,
        "rooms": len(sheets["Rooms"]) if "Rooms" in sheets else 0,
        "courses": len(sheets["Courses"]) if "Courses" in sheets else 0,
        "groups": len(sheets["Groups"]) if "Groups" in sheets else 0,
        "time_slots": len(sheets["TimeSlots"]) if "TimeSlots" in sheets else 0,
        "warnings": 0,
    }

    teachers_df = sheets.get("Teachers")
    courses_df = sheets.get("Courses")
    rooms_df = sheets.get("Rooms")
    groups_df = sheets.get("Groups")
    slots_df = sheets.get("TimeSlots")
    avail_df = sheets.get("TeacherAvailability")

    if teachers_df is not None and {"teacher_id", "teacher_name"}.issubset(teachers_df.columns):
        teacher_ids = set()
        for idx, row in teachers_df.iterrows():
            teacher_id = clean_cell(row.get("teacher_id"))
            if not teacher_id:
                errors.append(f"Teachers sheet row {idx + 2} is missing teacher_id.")
            elif teacher_id in teacher_ids:
                warnings.append(f"Duplicate teacher_id detected: {teacher_id}.")
            else:
                teacher_ids.add(teacher_id)
            if not clean_cell(row.get("teacher_name")):
                warnings.append(f"Teacher {teacher_id or 'row ' + str(idx + 2)} is missing teacher_name.")
    else:
        teacher_ids = set()

    if groups_df is not None and {"group_id", "student_count"}.issubset(groups_df.columns):
        group_ids = set()
        group_sizes = {}
        for idx, row in groups_df.iterrows():
            group_id = clean_cell(row.get("group_id"))
            size = parse_float(row.get("student_count"))
            if not group_id:
                errors.append(f"Groups sheet row {idx + 2} is missing group_id.")
                continue
            if group_id in group_ids:
                warnings.append(f"Duplicate group_id detected: {group_id}.")
            group_ids.add(group_id)
            if size is None or size <= 0:
                errors.append(f"Student count must be a number for group {group_id}.")
            else:
                group_sizes[group_id] = int(size)
    else:
        group_ids, group_sizes = set(), {}

    if rooms_df is not None and {"room_id", "room_name", "room_type", "capacity"}.issubset(rooms_df.columns):
        room_ids = set()
        room_records = []
        for idx, row in rooms_df.iterrows():
            room_id = clean_cell(row.get("room_id"))
            room_name = clean_cell(row.get("room_name")) or room_id or f"row {idx + 2}"
            capacity = parse_float(row.get("capacity"))
            room_type = clean_cell(row.get("room_type")).lower() or "normal"
            if not room_id:
                errors.append(f"Rooms sheet row {idx + 2} is missing room_id.")
            elif room_id in room_ids:
                warnings.append(f"Duplicate room_id detected: {room_id}.")
            else:
                room_ids.add(room_id)
            if capacity is None:
                errors.append(f"Room capacity must be a number. Check room {room_name}.")
            elif capacity <= 0:
                errors.append(f"Room capacity must be greater than zero. Check room {room_name}.")
            else:
                room_records.append({"id": room_id, "name": room_name, "type": room_type, "capacity": int(capacity)})
    else:
        room_records = []

    availability = get_available_windows(avail_df) if avail_df is not None else {}
    slots = get_timeslots(slots_df) if slots_df is not None else []

    if slots_df is not None and {"day", "start_time", "end_time"}.issubset(slots_df.columns):
        seen_slots = set()
        for idx, row in slots_df.iterrows():
            day = normalize_day(row.get("day"))
            start = parse_time_value(row.get("start_time"))
            end = parse_time_value(row.get("end_time"))
            if not day or start is None or end is None or start >= end:
                errors.append(f"TimeSlots sheet row {idx + 2} has an invalid time range.")
                continue
            if day not in WEEKDAYS:
                warnings.append(f"Time slot {clean_cell(row.get('slot_id')) or idx + 2} uses unsupported day {day}.")
            key = (day, start, end)
            if key in seen_slots:
                warnings.append(f"Duplicate time slot detected: {day} {format_time_minutes(start)}-{format_time_minutes(end)}.")
            seen_slots.add(key)

    if avail_df is not None and {"teacher_id", "day", "start_time", "end_time"}.issubset(avail_df.columns):
        for idx, row in avail_df.iterrows():
            teacher_id = clean_cell(row.get("teacher_id"))
            day = normalize_day(row.get("day"))
            start = parse_time_value(row.get("start_time"))
            end = parse_time_value(row.get("end_time"))
            if teacher_id and teacher_ids and teacher_id not in teacher_ids:
                errors.append(f"TeacherAvailability sheet uses teacher {teacher_id}, but it does not exist in Teachers sheet.")
            if not day or start is None or end is None or start >= end:
                errors.append(f"TeacherAvailability sheet row {idx + 2} has an invalid time range.")

    if courses_df is not None and {"course_id", "course_name", "teacher_id", "group_id", "course_type", "weekly_hours", "required_room_type"}.issubset(courses_df.columns):
        for idx, row in courses_df.iterrows():
            course_id = clean_cell(row.get("course_id")) or f"row {idx + 2}"
            teacher_id = clean_cell(row.get("teacher_id"))
            group_id = clean_cell(row.get("group_id"))
            course_type = clean_cell(row.get("course_type")).lower()
            required_room_type = clean_cell(row.get("required_room_type")).lower()
            weekly_hours = parse_float(row.get("weekly_hours"))

            if not clean_cell(row.get("course_name")):
                warnings.append(f"Course {course_id} is missing course_name.")
            if teacher_id and teacher_ids and teacher_id not in teacher_ids:
                errors.append(f"Teacher {teacher_id} is used in Courses sheet but does not exist in Teachers sheet.")
            if group_id and group_ids and group_id not in group_ids:
                errors.append(f"Group {group_id} is used in Courses sheet but does not exist in Groups sheet.")
            if not teacher_id:
                errors.append(f"Course {course_id} is missing teacher_id.")
            if not group_id:
                errors.append(f"Course {course_id} is missing group_id.")
            if weekly_hours is None:
                errors.append(f"Course {course_id} weekly_hours must be a number.")
            elif weekly_hours <= 0:
                errors.append(f"Course {course_id} weekly_hours must be greater than zero.")
            if teacher_id and teacher_id not in availability:
                warnings.append(f"Course {course_id} uses teacher {teacher_id}, but that teacher has no available time windows.")

            group_size = group_sizes.get(group_id, 0)
            needs_lab = required_room_type == "lab" or course_type == "lab"
            if room_records and group_size:
                compatible = [
                    room for room in room_records
                    if room["capacity"] >= group_size and (not needs_lab or room["type"] == "lab")
                ]
                if not compatible:
                    room_type_label = "lab" if needs_lab else "suitable"
                    warnings.append(f"Course {course_id} has no {room_type_label} room with capacity for group {group_id}.")

            if slots and teacher_id in availability:
                has_matching_slot = any(
                    any(window_start <= slot["start_min"] and slot["end_min"] <= window_end
                        for window_start, window_end in availability.get(teacher_id, {}).get(slot["day"], []))
                    for slot in slots
                )
                if not has_matching_slot:
                    warnings.append(f"Course {course_id} has no time slot inside teacher {teacher_id} availability.")

    summary["warnings"] = len(warnings)
    return {
        "summary": summary,
        "errors": sorted(set(errors)),
        "warnings": sorted(set(warnings)),
        "can_generate": len(errors) == 0,
    }

def intervals_overlap(start_a, end_a, start_b, end_b):
    return start_a < end_b and start_b < end_a

def resource_is_free(busy_map, resource_id, day, start_min, end_min):
    for booked_start, booked_end in busy_map.get((resource_id, day), []):
        if intervals_overlap(start_min, end_min, booked_start, booked_end):
            return False
    return True

def mark_resource_busy(busy_map, resource_id, day, start_min, end_min):
    busy_map.setdefault((resource_id, day), []).append((start_min, end_min))

def teacher_is_available(availability, teacher_id, day, start_min, end_min):
    return any(
        window_start <= start_min and end_min <= window_end
        for window_start, window_end in availability.get(teacher_id, {}).get(day, [])
    )

def load_generation_input(sheets):
    teachers_df = sheets["Teachers"]
    rooms_df = sheets["Rooms"]
    groups_df = sheets["Groups"]
    courses_df = sheets["Courses"]
    avail_df = sheets["TeacherAvailability"]
    slots_df = sheets["TimeSlots"]
    constraints_df = sheets.get("Constraints")

    teachers = {}
    for _, row in teachers_df.iterrows():
        teacher_id = clean_cell(row.get("teacher_id"))
        if not teacher_id:
            continue
        max_hours = parse_float(row.get("max_weekly_hours"), default=999)
        teachers[teacher_id] = {
            "id": teacher_id,
            "name": clean_cell(row.get("teacher_name")) or teacher_id,
            "max_weekly_hours": max_hours if max_hours and max_hours > 0 else 999,
        }

    rooms = []
    for _, row in rooms_df.iterrows():
        capacity = parse_float(row.get("capacity"))
        if capacity is None:
            continue
        rooms.append({
            "id": clean_cell(row.get("room_id")),
            "name": clean_cell(row.get("room_name")) or clean_cell(row.get("room_id")),
            "type": clean_cell(row.get("room_type")).lower() or "normal",
            "capacity": int(capacity),
        })

    groups = {}
    for _, row in groups_df.iterrows():
        group_id = clean_cell(row.get("group_id"))
        size = parse_float(row.get("student_count"), default=0)
        if group_id:
            groups[group_id] = {
                "id": group_id,
                "name": clean_cell(row.get("group_name")) or group_id,
                "size": int(size or 0),
            }

    constraints = {}
    if constraints_df is not None and {"constraint_name", "value"}.issubset(constraints_df.columns):
        for _, row in constraints_df.iterrows():
            name = clean_cell(row.get("constraint_name"))
            if name:
                constraints[name] = parse_bool(row.get("value"))

    availability = get_available_windows(avail_df)
    slots = get_timeslots(slots_df)
    average_slot_hours = sum(slot["duration_hours"] for slot in slots) / len(slots) if slots else 1.5

    sessions = []
    for _, row in courses_df.iterrows():
        course_id = clean_cell(row.get("course_id"))
        teacher_id = clean_cell(row.get("teacher_id"))
        group_id = clean_cell(row.get("group_id"))
        course_type = clean_cell(row.get("course_type")).lower() or "lecture"
        required_room_type = clean_cell(row.get("required_room_type")).lower() or ("lab" if course_type == "lab" else "normal")
        weekly_hours = parse_float(row.get("weekly_hours"), default=0) or 0
        session_count = max(1, int(math.ceil(weekly_hours / average_slot_hours))) if average_slot_hours > 0 else 1
        teacher = teachers.get(teacher_id, {"name": teacher_id, "max_weekly_hours": 999})
        group = groups.get(group_id, {"name": group_id, "size": 0})
        for session_index in range(1, session_count + 1):
            sessions.append({
                "course_id": course_id,
                "course_name": clean_cell(row.get("course_name")) or course_id,
                "teacher_id": teacher_id,
                "teacher_name": teacher["name"],
                "group_id": group_id,
                "group_name": group["name"],
                "group_size": group["size"],
                "course_type": course_type,
                "required_room_type": required_room_type,
                "weekly_hours": weekly_hours,
                "session_index": session_index,
                "session_count": session_count,
            })

    return teachers, rooms, groups, availability, slots, constraints, sessions

def explain_unresolved_session(session, rooms, availability, slots, teacher_busy, group_busy, room_busy):
    needs_lab = session["required_room_type"] == "lab" or session["course_type"] == "lab"
    capacity_rooms = [room for room in rooms if room["capacity"] >= session["group_size"]]
    if needs_lab and not any(room["type"] == "lab" and room["capacity"] >= session["group_size"] for room in rooms):
        return "No available lab room has enough capacity for this group."
    if not capacity_rooms:
        return "No room has enough capacity for this group."
    if not any(teacher_is_available(availability, session["teacher_id"], slot["day"], slot["start_min"], slot["end_min"]) for slot in slots):
        return "No time slot falls inside this teacher's availability window."
    if all(not resource_is_free(teacher_busy, session["teacher_id"], slot["day"], slot["start_min"], slot["end_min"]) for slot in slots):
        return "The assigned teacher is already booked in every compatible time slot."
    if all(not resource_is_free(group_busy, session["group_id"], slot["day"], slot["start_min"], slot["end_min"]) for slot in slots):
        return "The student group is already booked in every compatible time slot."
    if all(
        not any(resource_is_free(room_busy, room["id"], slot["day"], slot["start_min"], slot["end_min"]) for room in capacity_rooms)
        for slot in slots
    ):
        return "All rooms with enough capacity are already booked in compatible time slots."
    return "No valid combination satisfies teacher, group, room, availability, and capacity constraints."

def generate_optimized_timetable(sheets):
    teachers, rooms, groups, availability, slots, constraints, sessions = load_generation_input(sheets)
    if not slots:
        return [], [
            {"course_id": "TimeSlots", "course_name": "TimeSlots", "reason": "No valid time slots were found."}
        ], 0

    def matching_room_count(session):
        needs_lab = session["required_room_type"] == "lab" or session["course_type"] == "lab"
        return len([
            room for room in rooms
            if room["capacity"] >= session["group_size"] and (not needs_lab or room["type"] == "lab")
        ])

    sessions.sort(key=lambda item: (
        0 if (item["required_room_type"] == "lab" or item["course_type"] == "lab") else 1,
        matching_room_count(item),
        -item["group_size"],
        item["teacher_id"],
        item["course_id"],
    ))

    teacher_busy, room_busy, group_busy = {}, {}, {}
    teacher_hours, teacher_day_count = {}, {}
    group_day_count, course_day_count = {}, {}
    timetable, conflicts = [], []
    score_total = 0

    def candidate_rooms(session):
        needs_lab = session["required_room_type"] == "lab" or session["course_type"] == "lab"
        fitting = [room for room in rooms if room["capacity"] >= session["group_size"]]
        if needs_lab:
            return [room for room in fitting if room["type"] == "lab"]
        preferred = [room for room in fitting if room["type"] == "normal"]
        return preferred if preferred else fitting

    def group_gap_score(group_id, day, start_min, end_min):
        existing = group_busy.get((group_id, day), [])
        if not existing:
            return 6
        nearest_gap = min(
            min(abs(start_min - booked_end), abs(booked_start - end_min))
            for booked_start, booked_end in existing
        )
        if nearest_gap <= 30:
            return 10
        if nearest_gap <= 120:
            return 3
        return -6

    for session_item in sessions:
        candidates = []
        teacher_id = session_item["teacher_id"]
        group_id = session_item["group_id"]
        teacher = teachers.get(teacher_id, {"max_weekly_hours": 999})
        rooms_for_session = candidate_rooms(session_item)

        for slot in slots:
            day = slot["day"]
            start_min = slot["start_min"]
            end_min = slot["end_min"]
            duration = slot["duration_hours"]
            if constraints.get("respect_teacher_availability", True) and not teacher_is_available(availability, teacher_id, day, start_min, end_min):
                continue
            if not resource_is_free(teacher_busy, teacher_id, day, start_min, end_min):
                continue
            if not resource_is_free(group_busy, group_id, day, start_min, end_min):
                continue
            projected_teacher_hours = teacher_hours.get(teacher_id, 0) + duration
            if projected_teacher_hours > teacher.get("max_weekly_hours", 999) + 0.01:
                continue

            for room in rooms_for_session:
                if not resource_is_free(room_busy, room["id"], day, start_min, end_min):
                    continue
                score = 100
                if constraints.get("balance_daily_workload", True):
                    score -= group_day_count.get((group_id, day), 0) * 14
                    score -= teacher_day_count.get((teacher_id, day), 0) * 7
                if constraints.get("avoid_late_classes", True):
                    if start_min < 12 * 60:
                        score += 12
                    elif start_min < 15 * 60:
                        score += 6
                    else:
                        score -= 12
                score += group_gap_score(group_id, day, start_min, end_min)
                if course_day_count.get((session_item["course_id"], group_id, day), 0):
                    score -= 16
                room_waste = max(0, room["capacity"] - session_item["group_size"])
                score -= min(room_waste, 40) * 0.15
                if (session_item["required_room_type"] == "normal" and room["type"] == "normal") or (session_item["required_room_type"] == "lab" and room["type"] == "lab"):
                    score += 5
                score -= teacher_hours.get(teacher_id, 0) * 0.4
                candidates.append((score, slot, room))

        if not candidates:
            conflicts.append({
                "course_id": session_item["course_id"],
                "course_name": session_item["course_name"],
                "reason": explain_unresolved_session(session_item, rooms, availability, slots, teacher_busy, group_busy, room_busy),
            })
            continue

        candidates.sort(key=lambda item: (-item[0], WEEKDAYS.index(item[1]["day"]) if item[1]["day"] in WEEKDAYS else 99, item[1]["start_min"], item[2]["capacity"], item[2]["id"]))
        best_score, slot, room = candidates[0]
        day = slot["day"]
        start_min = slot["start_min"]
        end_min = slot["end_min"]
        duration = slot["duration_hours"]

        mark_resource_busy(teacher_busy, teacher_id, day, start_min, end_min)
        mark_resource_busy(room_busy, room["id"], day, start_min, end_min)
        mark_resource_busy(group_busy, group_id, day, start_min, end_min)
        teacher_hours[teacher_id] = teacher_hours.get(teacher_id, 0) + duration
        teacher_day_count[(teacher_id, day)] = teacher_day_count.get((teacher_id, day), 0) + 1
        group_day_count[(group_id, day)] = group_day_count.get((group_id, day), 0) + 1
        course_day_count[(session_item["course_id"], group_id, day)] = course_day_count.get((session_item["course_id"], group_id, day), 0) + 1
        score_total += max(0, min(120, best_score))

        timetable.append({
            "day": day,
            "start_time": slot["start"],
            "end_time": slot["end"],
            "course_id": session_item["course_id"],
            "course_name": session_item["course_name"],
            "teacher_name": session_item["teacher_name"],
            "room_name": room["name"],
            "group_name": session_item["group_name"],
            "course_type": session_item["course_type"],
            "status": "scheduled",
        })

    scheduled = len(timetable)
    total = len(sessions)
    completion_score = (scheduled / total) * 85 if total else 0
    placement_score = ((score_total / scheduled) / 120) * 15 if scheduled else 0
    optimization_score = round(min(100, completion_score + placement_score), 1)
    return timetable, conflicts, optimization_score

@app.route("/admin/download-template")
@role_required("admin")
def download_template():
    if not EXCEL_AVAILABLE:
        flash("pandas/openpyxl not installed.", "error")
        return redirect(url_for("timetable_builder"))
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="0F8F5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    def make_sheet(wb, name, headers, rows, first=False):
        ws = wb.active if first else wb.create_sheet(name)
        if first: ws.title = name
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for row in rows: ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4
    make_sheet(wb, "Teachers",
        ["teacher_id","teacher_name","department","email","max_weekly_hours"],
        [["T001","Nigar Alishzada","Computer Science","nigar@karabakh.edu.az",16],
         ["T002","Elvin Karimov","Mathematics","elvin@karabakh.edu.az",14],
         ["T003","Rauf Mammadov","Computer Science","rauf@karabakh.edu.az",18]], first=True)
    make_sheet(wb, "TeacherAvailability",
        ["teacher_id","day","start_time","end_time","availability_status"],
        [["T001","Monday","09:00","17:00","available"],
         ["T001","Tuesday","09:00","17:00","available"],
         ["T001","Wednesday","09:00","17:00","available"],
         ["T001","Thursday","09:00","17:00","available"],
         ["T001","Friday","09:00","15:00","available"],
         ["T002","Monday","09:00","15:00","available"],
         ["T002","Tuesday","10:00","16:00","available"],
         ["T002","Wednesday","09:00","17:00","available"],
         ["T002","Thursday","09:00","17:00","available"],
         ["T003","Monday","09:00","17:00","available"],
         ["T003","Tuesday","09:00","17:00","available"],
         ["T003","Wednesday","09:00","15:00","available"],
         ["T003","Friday","09:00","17:00","available"]])
    make_sheet(wb, "Courses",
        ["course_id","course_name","teacher_id","group_id","course_type","weekly_hours","required_room_type"],
        [["C001","Web Engineering","T001","G001","lecture",3,"normal"],
         ["C002","Artificial Intelligence","T001","G001","lecture",4,"normal"],
         ["C003","Discrete Mathematics","T002","G002","lecture",3,"normal"],
         ["C004","Computer Graphics","T003","G001","lab",2,"lab"],
         ["C005","Data Analysis","T003","G002","lab",2,"lab"],
         ["C006","Database Systems","T002","G003","lecture",3,"normal"],
         ["C007","Probability & Statistics","T002","G003","lecture",3,"normal"],
         ["C008","Network Programming","T001","G002","lecture",2,"normal"]])
    make_sheet(wb, "Rooms",
        ["room_id","room_name","room_type","capacity"],
        [["R001","A-101","normal",40],["R002","A-204","normal",35],
         ["R003","B-118","normal",30],["R004","C-301","normal",50],
         ["R005","Lab-1","lab",32],["R006","Lab-2","lab",30],["R007","Lab-3","lab",28]])
    make_sheet(wb, "Groups",
        ["group_id","group_name","student_count","department","semester"],
        [["G001","KOM24A",28,"Computer Science",4],
         ["G002","KOM24B",26,"Computer Science",4],
         ["G003","KOM25A",30,"Computer Science",2]])
    make_sheet(wb, "TimeSlots",
        ["slot_id","day","start_time","end_time"],
        [["S001","Monday","09:00","10:30"],["S002","Monday","10:40","12:10"],
         ["S003","Monday","13:00","14:30"],["S004","Monday","14:40","16:10"],
         ["S005","Tuesday","09:00","10:30"],["S006","Tuesday","10:40","12:10"],
         ["S007","Tuesday","13:00","14:30"],["S008","Tuesday","14:40","16:10"],
         ["S009","Wednesday","09:00","10:30"],["S010","Wednesday","10:40","12:10"],
         ["S011","Wednesday","13:00","14:30"],["S012","Wednesday","14:40","16:10"],
         ["S013","Thursday","09:00","10:30"],["S014","Thursday","10:40","12:10"],
         ["S015","Thursday","13:00","14:30"],["S016","Thursday","14:40","16:10"],
         ["S017","Friday","09:00","10:30"],["S018","Friday","10:40","12:10"],
         ["S019","Friday","13:00","14:30"]])
    make_sheet(wb, "Constraints",
        ["constraint_name","constraint_type","priority","value"],
        [["no_teacher_overlap","hard","high","true"],
         ["no_room_overlap","hard","high","true"],
         ["no_group_overlap","hard","high","true"],
         ["match_room_type","hard","high","true"],
         ["respect_teacher_availability","hard","high","true"],
         ["balance_daily_workload","soft","medium","true"],
         ["avoid_late_classes","soft","low","true"]])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="smart_timetable_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/upload-excel", methods=["POST"])
@role_required("admin")
def upload_excel():
    if not EXCEL_AVAILABLE:
        return jsonify({"success": False, "error": "Excel libraries not installed. Run: pip install pandas openpyxl"})
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file provided."})
    f = request.files['file']
    if f.filename == '':
        return jsonify({"success": False, "error": "No file selected."})
    if not allowed_file(f.filename):
        return jsonify({"success": False, "error": "Only .xlsx and .xls files are allowed."})
    original = secure_filename(f.filename)
    if not original:
        return jsonify({"success": False, "error": "Invalid file name."})
    name_root, ext = os.path.splitext(original)
    filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}_{name_root}{ext}"
    filename, filepath = safe_upload_path(filename)
    if not filepath:
        return jsonify({"success": False, "error": "Invalid upload path."})
    f.save(filepath)
    clear_generated_timetable()
    rec = UploadedTimetableFile(filename=filename, uploaded_by=session["user_id"], analysis_status="uploaded")
    db.session.add(rec)
    db.session.commit()
    return jsonify({
        "success": True,
        "filename": filename,
        "upload_id": rec.id,
        "message": "New Excel file uploaded. Previous generated draft was cleared.",
    })

@app.route("/admin/analyze-excel", methods=["POST"])
@role_required("admin")
def analyze_excel():
    if not EXCEL_AVAILABLE:
        return jsonify({"success": False, "error": "Excel libraries not installed."})
    data = request.get_json() or {}
    filename = data.get("filename", "")
    filename, filepath = safe_upload_path(filename)
    if not filename or not filepath:
        return jsonify({"success": False, "error": "Invalid filename."})
    if not os.path.exists(filepath):
        return jsonify({"success": False, "error": "File not found on server."})
    clear_generated_timetable()
    db.session.commit()
    try:
        sheets, schema_errors = read_excel_sheets(filepath)
        result = validate_excel_data(sheets, schema_errors)
        UploadedTimetableFile.query.filter_by(filename=filename).update({"analysis_status": "analyzed" if result["can_generate"] else "issues"})
        db.session.commit()
        return jsonify({
            "success": True,
            "summary": result["summary"],
            "errors": result["errors"],
            "warnings": result["warnings"],
            "can_generate": result["can_generate"],
        })
    except ValueError as e:
        db.session.rollback()
        UploadedTimetableFile.query.filter_by(filename=filename).update({"analysis_status": "issues"})
        db.session.commit()
        return jsonify({"success": False, "error": str(e)})
    except Exception:
        db.session.rollback()
        UploadedTimetableFile.query.filter_by(filename=filename).update({"analysis_status": "issues"})
        db.session.commit()
        return jsonify({"success": False, "error": "Could not read Excel file. Please check the workbook format and required sheets."})

@app.route("/admin/generate-from-excel", methods=["POST"])
@role_required("admin")
def generate_from_excel():
    if not EXCEL_AVAILABLE:
        return jsonify({"success": False, "error": "Excel libraries not installed."})
    data = request.get_json() or {}
    filename = data.get("filename", "")
    filename, filepath = safe_upload_path(filename)
    if not filename or not filepath:
        return jsonify({"success": False, "error": "Invalid filename."})
    if not os.path.exists(filepath):
        return jsonify({"success": False, "error": "File not found."})
    clear_generated_timetable()
    db.session.commit()
    try:
        sheets, schema_errors = read_excel_sheets(filepath)
        validation = validate_excel_data(sheets, schema_errors)
        if not validation["can_generate"]:
            UploadedTimetableFile.query.filter_by(filename=filename).update({"analysis_status": "issues"})
            db.session.commit()
            return jsonify({
                "success": False,
                "error": "Excel validation failed. Fix the listed issues before generating.",
                "errors": validation["errors"],
                "warnings": validation["warnings"],
            })

        timetable, conflicts, optimization_score = generate_optimized_timetable(sheets)
        for entry in timetable:
            db.session.add(GeneratedTimetable(**entry))
        UploadedTimetableFile.query.filter_by(filename=filename).update({"analysis_status": "generated"})
        settings = get_or_create_settings()
        settings.is_timetable_published = False
        db.session.commit()

        total = len(timetable) + len(conflicts)
        scheduled = len(timetable)
        return jsonify({
            "success": True,
            "summary": {
                "total_sessions": total,
                "scheduled_sessions": scheduled,
                "unresolved_sessions": len(conflicts),
                "conflict_count": len(conflicts),
                "optimization_score": optimization_score,
            },
            "timetable": timetable,
            "conflicts": conflicts,
            "warnings": validation["warnings"],
        })
    except ValueError as e:
        db.session.rollback()
        UploadedTimetableFile.query.filter_by(filename=filename).update({"analysis_status": "issues"})
        db.session.commit()
        return jsonify({"success": False, "error": str(e)})
    except Exception:
        db.session.rollback()
        UploadedTimetableFile.query.filter_by(filename=filename).update({"analysis_status": "issues"})
        db.session.commit()
        return jsonify({"success": False, "error": "Generation failed. Please verify the Excel data and try again."})

@app.route("/admin/export-timetable")
@role_required("admin")
def export_timetable():
    if not EXCEL_AVAILABLE:
        flash("Excel libraries not installed.", "error")
        return redirect(url_for("timetable_builder"))
    entries = GeneratedTimetable.query.order_by(GeneratedTimetable.day, GeneratedTimetable.start_time).all()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Generated Timetable"
    hdr_fill = PatternFill("solid", fgColor="0F8F5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = ["Day","Start Time","End Time","Course ID","Course Name","Teacher","Room","Group","Type","Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for e in entries:
        ws.append([e.day, e.start_time, e.end_time, e.course_id, e.course_name,
                   e.teacher_name, e.room_name, e.group_name, e.course_type, e.status])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="generated_timetable.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────── HELPERS ───────────────
def safe_text(value):
    if value is None:
        return ""
    return str(value).strip()

def unique_by_id(items):
    seen = set()
    unique = []
    for item in items:
        if not item or item.id in seen:
            continue
        seen.add(item.id)
        unique.append(item)
    return unique

def build_teacher_timetable_context(user, extra_lesson_form=None, available_rooms=None):
    is_published = get_publish_status()
    if is_published:
        schedule = ScheduleEntry.query.filter_by(teacher_id=user.id).order_by(ScheduleEntry.day, ScheduleEntry.time_start).all()
    else:
        schedule = []
    entries = build_timetable_context(schedule)
    days_order = list(WEEKDAYS)
    timetable_by_day = {day: [] for day in days_order}
    for item in entries:
        day = item["entry"].day
        if day not in timetable_by_day:
            timetable_by_day[day] = []
            days_order.append(day)
        timetable_by_day[day].append(item)

    teacher_courses = Course.query.filter_by(teacher_id=user.id).order_by(Course.code, Course.name).all()
    extra_lesson_requests = build_extra_lesson_request_rows(
        ExtraLessonRequest.query.filter_by(teacher_id=user.id).order_by(ExtraLessonRequest.created_at.desc()).limit(5).all()
    )
    form_data = extra_lesson_form or {
        "course_id": str(teacher_courses[0].id) if teacher_courses else "",
        "date": "",
        "start_time": "",
        "end_time": "",
        "message": "",
    }
    return {
        "user": user,
        "timetable_by_day": timetable_by_day,
        "days_order": days_order,
        "is_published": is_published,
        "total": len(entries),
        "teacher_courses": teacher_courses,
        "extra_lesson_form": form_data,
        "available_rooms": available_rooms,
        "extra_lesson_requests": extra_lesson_requests,
    }

def parse_extra_lesson_form(user, require_room=False):
    errors = []
    course_id = request.form.get("course_id", type=int)
    date_text = safe_text(request.form.get("date"))
    start_text = safe_text(request.form.get("start_time"))
    end_text = safe_text(request.form.get("end_time"))
    message = safe_text(request.form.get("message"))[:500]
    room_id = request.form.get("room_id", type=int)

    course = Course.query.filter_by(id=course_id, teacher_id=user.id).first() if course_id else None
    if not course:
        errors.append("Please select one of your courses.")

    lesson_date = None
    day = ""
    if date_text:
        try:
            lesson_date = datetime.strptime(date_text, "%Y-%m-%d").date()
            day = lesson_date.strftime("%A")
        except ValueError:
            errors.append("Please enter a valid lesson date.")
    else:
        errors.append("Please select a lesson date.")

    start_min = parse_time_value(start_text)
    end_min = parse_time_value(end_text)
    if start_min is None or end_min is None or start_min >= end_min:
        errors.append("Please enter a valid start and end time.")

    if require_room and not room_id:
        errors.append("Please select an available room.")

    return {
        "course_id": str(course_id or ""),
        "course": course,
        "room_id": room_id,
        "date": date_text,
        "lesson_date": lesson_date,
        "day": day,
        "start_time": format_time_minutes(start_min) if start_min is not None else start_text,
        "end_time": format_time_minutes(end_min) if end_min is not None else end_text,
        "start_min": start_min,
        "end_min": end_min,
        "message": message,
    }, errors

def find_available_extra_lesson_rooms(course, day, lesson_date, start_min, end_min):
    room_query = Room.query
    if course and course.is_lab:
        room_query = room_query.filter_by(is_lab=True)
    rooms = room_query.order_by(Room.name).all()
    return find_empty_rooms_for_slot(rooms, day, start_min, end_min, lesson_date=lesson_date)

def teacher_has_slot_conflict(user, day, lesson_date, start_min, end_min):
    for entry in ScheduleEntry.query.filter_by(teacher_id=user.id, day=day).all():
        entry_start = parse_time_value(entry.time_start)
        entry_end = parse_time_value(entry.time_end)
        if entry_start is not None and entry_end is not None and intervals_overlap(entry_start, entry_end, start_min, end_min):
            return True

    for row in GeneratedTimetable.query.filter_by(day=day).all():
        if normalize_lookup_text(row.status) and normalize_lookup_text(row.status) != "scheduled":
            continue
        if normalize_lookup_text(row.teacher_name) != normalize_lookup_text(user.full_name):
            continue
        row_start = parse_time_value(row.start_time)
        row_end = parse_time_value(row.end_time)
        if row_start is not None and row_end is not None and intervals_overlap(row_start, row_end, start_min, end_min):
            return True

    for request_row in ExtraLessonRequest.query.filter_by(teacher_id=user.id).all():
        if normalize_lookup_text(request_row.status) == "cancelled":
            continue
        if not dates_match(request_row.lesson_date, lesson_date):
            continue
        row_start = parse_time_value(request_row.start_time)
        row_end = parse_time_value(request_row.end_time)
        if row_start is not None and row_end is not None and intervals_overlap(row_start, row_end, start_min, end_min):
            return True
    return False

def build_extra_lesson_request_rows(requests):
    rows = []
    for request_row in requests:
        teacher = User.query.get(request_row.teacher_id)
        course = Course.query.get(request_row.course_id)
        room = Room.query.get(request_row.room_id)
        teacher_name = safe_text(teacher.full_name if teacher else "Unknown teacher")
        course_label = safe_text(f"{course.code} - {course.name}" if course else "Unknown course")
        room_name = safe_text(room.name if room else "Unknown room")
        lesson_date = format_lesson_date(request_row.lesson_date)
        admin_message = (
            f"{teacher_name} will teach an extra lesson for {course_label} in room {room_name} "
            f"on {lesson_date} from {safe_text(request_row.start_time)} to {safe_text(request_row.end_time)}."
        )
        if safe_text(request_row.message):
            admin_message = f"{admin_message} Note: {safe_text(request_row.message)}"
        rows.append({
            "request": request_row,
            "teacher": teacher,
            "course": course,
            "room": room,
            "lesson_date": lesson_date,
            "teacher_name": teacher_name,
            "course_label": course_label,
            "room_name": room_name,
            "admin_message": admin_message,
        })
    return rows

def build_extra_lesson_announcement(request_row):
    row = build_extra_lesson_request_rows([request_row])[0]
    return Announcement(
        title=row["admin_message"],
        announcement_type="important",
        audience="all",
        source_type="extra_lesson",
        source_id=request_row.id,
    )

def build_recent_announcements(audience, limit=10):
    allowed = ["all", audience]
    rows = Announcement.query.filter(Announcement.audience.in_(allowed)).order_by(Announcement.created_at.desc()).limit(limit).all()
    return [
        {
            "title": safe_text(row.title),
            "date": format_lesson_date(row.created_at),
            "type": safe_text(row.announcement_type) or "info",
        }
        for row in rows
    ]

def format_lesson_date(value):
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return safe_text(value)

def dates_match(stored_date, selected_date):
    if selected_date is None:
        return False
    if hasattr(stored_date, "isoformat"):
        return stored_date == selected_date
    return safe_text(stored_date) == selected_date.isoformat()

def build_admin_student_rows(students):
    rows = []
    for student in students:
        group = Group.query.get(student.group_id) if student.group_id else None
        courses = []
        for enrollment in Enrollment.query.filter_by(student_id=student.id).all():
            course = Course.query.get(enrollment.course_id)
            teacher = User.query.get(course.teacher_id) if course else None
            if course:
                courses.append({"course": course, "teacher": teacher, "enrollment": enrollment})
        rows.append({"student": student, "group": group, "courses": courses})
    return rows

def build_admin_teacher_rows(teachers):
    rows = []
    for teacher in teachers:
        courses = Course.query.filter_by(teacher_id=teacher.id).order_by(Course.code, Course.name).all()
        schedule = []
        for entry in ScheduleEntry.query.filter_by(teacher_id=teacher.id).order_by(ScheduleEntry.day, ScheduleEntry.time_start).all():
            schedule.append({
                "entry": entry,
                "course": Course.query.get(entry.course_id),
                "room": Room.query.get(entry.room_id),
                "group": Group.query.get(entry.group_id),
            })
        student_ids = set()
        for course in courses:
            for enrollment in Enrollment.query.filter_by(course_id=course.id).all():
                student_ids.add(enrollment.student_id)
        rows.append({
            "teacher": teacher,
            "courses": courses,
            "schedule": schedule,
            "student_count": len(student_ids),
        })
    return rows

def students_for_course_group(course_id, group_id):
    enrolled_students = []
    for enrollment in Enrollment.query.filter_by(course_id=course_id).all():
        student = User.query.get(enrollment.student_id)
        if student and student.group_id == group_id:
            enrolled_students.append(student)
    if enrolled_students:
        return sorted(unique_by_id(enrolled_students), key=lambda item: item.full_name)
    return User.query.filter_by(role="student", group_id=group_id).order_by(User.full_name).all()

def build_admin_course_rows(courses, t_map):
    rows = []
    for course in courses:
        groups = []
        for entry in ScheduleEntry.query.filter_by(course_id=course.id).all():
            group = Group.query.get(entry.group_id)
            if group:
                groups.append(group)
        for enrollment in Enrollment.query.filter_by(course_id=course.id).all():
            student = User.query.get(enrollment.student_id)
            group = Group.query.get(student.group_id) if student and student.group_id else None
            if group:
                groups.append(group)

        group_rows = []
        for group in sorted(unique_by_id(groups), key=lambda item: item.name):
            group_rows.append({
                "group": group,
                "students": students_for_course_group(course.id, group.id),
            })

        rows.append({
            "course": course,
            "teacher": t_map.get(course.teacher_id),
            "groups": group_rows,
        })
    return rows

def build_admin_group_rows(groups):
    rows = []
    for group in groups:
        students = User.query.filter_by(role="student", group_id=group.id).order_by(User.full_name).all()
        schedule_entries = ScheduleEntry.query.filter_by(group_id=group.id).order_by(ScheduleEntry.day, ScheduleEntry.time_start).all()
        courses = unique_by_id([Course.query.get(entry.course_id) for entry in schedule_entries])
        rows.append({"group": group, "students": students, "courses": courses})
    return rows

def generated_row_overlaps_room(row, room_name, day, start_min, end_min):
    if normalize_lookup_text(row.status) and normalize_lookup_text(row.status) != "scheduled":
        return False
    if normalize_lookup_text(row.room_name) != normalize_lookup_text(room_name):
        return False
    row_start = parse_time_value(row.start_time)
    row_end = parse_time_value(row.end_time)
    return row.day == day and row_start is not None and row_end is not None and intervals_overlap(row_start, row_end, start_min, end_min)

def schedule_entry_overlaps_room(entry, room_id, day, start_min, end_min):
    row_start = parse_time_value(entry.time_start)
    row_end = parse_time_value(entry.time_end)
    return entry.room_id == room_id and entry.day == day and row_start is not None and row_end is not None and intervals_overlap(row_start, row_end, start_min, end_min)

def extra_lesson_overlaps_room(request_row, room_id, day, start_min, end_min, lesson_date=None):
    if normalize_lookup_text(request_row.status) == "cancelled":
        return False
    if request_row.room_id != room_id:
        return False
    if lesson_date is not None:
        if not dates_match(request_row.lesson_date, lesson_date):
            return False
    elif request_row.day != day:
        return False
    row_start = parse_time_value(request_row.start_time)
    row_end = parse_time_value(request_row.end_time)
    return row_start is not None and row_end is not None and intervals_overlap(row_start, row_end, start_min, end_min)

def find_empty_rooms_for_slot(rooms, day, start_min, end_min, lesson_date=None):
    generated_rows = GeneratedTimetable.query.all()
    schedule_entries = ScheduleEntry.query.all()
    extra_lesson_requests = ExtraLessonRequest.query.all()
    results = []
    for room in rooms:
        if generated_rows:
            booked = any(generated_row_overlaps_room(row, room.name, day, start_min, end_min) for row in generated_rows)
        else:
            booked = any(schedule_entry_overlaps_room(entry, room.id, day, start_min, end_min) for entry in schedule_entries)
        if not booked:
            booked = any(extra_lesson_overlaps_room(row, room.id, day, start_min, end_min, lesson_date=lesson_date) for row in extra_lesson_requests)
        if not booked:
            results.append({
                "id": room.id,
                "name": room.name,
                "capacity": room.capacity,
                "type": "Laboratory" if room.is_lab else "Standard Class",
            })
    return results

def normalize_lookup_text(value):
    return " ".join(safe_text(value).lower().split())

def get_or_create_settings():
    settings = SystemSettings.query.first()
    if not settings:
        settings = SystemSettings(is_timetable_published=False)
        db.session.add(settings)
        db.session.flush()
    return settings

def get_default_university():
    uni = University.query.first()
    if not uni:
        uni = University(
            name="EMM",
            domain=f"generated-{uuid.uuid4().hex[:8]}.local",
            address="Imported timetable",
            founded=datetime.utcnow().year,
        )
        db.session.add(uni)
        db.session.flush()
    return uni

def initials_from_name(name):
    parts = [part for part in safe_text(name).split() if part]
    initials = "".join(part[0] for part in parts[:2]).upper()
    return initials[:5] or "T"

def generated_email_for_name(name, role):
    base = re.sub(r"[^a-z0-9]+", ".", safe_text(name).lower()).strip(".")
    base = base or uuid.uuid4().hex[:8]
    candidate = f"{role}-{base}@generated.local"
    suffix = 2
    while User.query.filter_by(email=candidate).first():
        candidate = f"{role}-{base}-{suffix}@generated.local"
        suffix += 1
    return candidate

def find_teacher_by_name(name):
    target = normalize_lookup_text(name)
    if not target:
        return None
    for teacher in User.query.filter_by(role="teacher").all():
        if normalize_lookup_text(teacher.full_name) == target:
            return teacher
    return None

def find_room_by_name(name):
    target = normalize_lookup_text(name)
    if not target:
        return None
    for room in Room.query.all():
        if normalize_lookup_text(room.name) == target:
            return room
    return None

def find_group_by_name(name):
    target = normalize_lookup_text(name)
    if not target:
        return None
    for group in Group.query.all():
        if normalize_lookup_text(group.name) == target:
            return group
    return None

def ensure_teacher_for_generated(row, university):
    name = safe_text(row.teacher_name) or "Unassigned Teacher"
    teacher = find_teacher_by_name(name)
    if teacher:
        return teacher
    teacher = User(
        full_name=name,
        email=generated_email_for_name(name, "teacher"),
        password_hash=generate_password_hash("password123"),
        role="teacher",
        university_id=university.id,
        department="Imported",
        avatar_initials=initials_from_name(name),
    )
    db.session.add(teacher)
    db.session.flush()
    return teacher

def ensure_room_for_generated(row, university):
    name = safe_text(row.room_name) or "Unassigned Room"
    room = find_room_by_name(name)
    if room:
        room.is_lab = room.is_lab or normalize_lookup_text(row.course_type) == "lab"
        return room
    room = Room(
        name=name,
        capacity=30,
        is_lab=normalize_lookup_text(row.course_type) == "lab",
        university_id=university.id,
    )
    db.session.add(room)
    db.session.flush()
    return room

def ensure_group_for_generated(row, university):
    name = safe_text(row.group_name) or "Unassigned Group"
    group = find_group_by_name(name)
    if group:
        return group
    group = Group(name=name, year=1, size=0, university_id=university.id)
    db.session.add(group)
    db.session.flush()
    return group

def ensure_course_for_generated(row, teacher, university):
    code = safe_text(row.course_id) or f"GEN{row.id}"
    name = safe_text(row.course_name) or code
    is_lab = normalize_lookup_text(row.course_type) == "lab"
    course = Course.query.filter_by(code=code).first()
    if not course:
        course = Course(
            code=code,
            name=name,
            credits=3,
            is_lab=is_lab,
            teacher_id=teacher.id,
            university_id=university.id,
            color="#0f8f5f" if not is_lab else "#7c3aed",
        )
        db.session.add(course)
        db.session.flush()
        return course
    course.name = name
    course.is_lab = is_lab
    course.teacher_id = teacher.id
    if not course.university_id:
        course.university_id = university.id
    return course

def sync_generated_timetable_to_schedule():
    generated_rows = GeneratedTimetable.query.order_by(GeneratedTimetable.day, GeneratedTimetable.start_time).all()
    if not generated_rows:
        return 0

    university = get_default_university()
    ScheduleEntry.query.delete()
    sent_count = 0

    for row in generated_rows:
        status = normalize_lookup_text(row.status) or "scheduled"
        if status != "scheduled":
            continue
        teacher = ensure_teacher_for_generated(row, university)
        room = ensure_room_for_generated(row, university)
        group = ensure_group_for_generated(row, university)
        course = ensure_course_for_generated(row, teacher, university)
        db.session.add(ScheduleEntry(
            course_id=course.id,
            teacher_id=teacher.id,
            room_id=room.id,
            group_id=group.id,
            day=safe_text(row.day),
            time_start=safe_text(row.start_time),
            time_end=safe_text(row.end_time),
            semester="Generated Timetable",
        ))
        sent_count += 1

    settings = get_or_create_settings()
    settings.is_timetable_published = sent_count > 0
    db.session.commit()
    return sent_count

def sort_master_rows(rows):
    day_order = {day: idx for idx, day in enumerate(WEEKDAYS)}
    return sorted(rows, key=lambda row: (
        day_order.get(row["day"], 99),
        row["time_start"],
        row["time_end"],
        row["course_code"],
        row["group_name"],
    ))

def build_generated_master_rows(generated_rows):
    rows = []
    for row in generated_rows:
        is_lab = normalize_lookup_text(row.course_type) == "lab"
        rows.append({
            "day": safe_text(row.day),
            "time_start": safe_text(row.start_time),
            "time_end": safe_text(row.end_time),
            "course_code": safe_text(row.course_id),
            "course_name": safe_text(row.course_name),
            "teacher_name": safe_text(row.teacher_name),
            "room_name": safe_text(row.room_name),
            "group_name": safe_text(row.group_name),
            "course_type": safe_text(row.course_type) or "lecture",
            "status": safe_text(row.status) or "scheduled",
            "is_lab": is_lab,
            "course_color": "#7c3aed" if is_lab else "#0f8f5f",
        })
    return sort_master_rows(rows)

def build_schedule_master_rows(schedule):
    rows = []
    for item in build_timetable_context(schedule):
        entry = item["entry"]
        course = item["course"]
        teacher = item["teacher"]
        room = item["room"]
        group = item["group"]
        is_lab = bool(course and course.is_lab)
        rows.append({
            "day": safe_text(entry.day),
            "time_start": safe_text(entry.time_start),
            "time_end": safe_text(entry.time_end),
            "course_code": safe_text(course.code if course else ""),
            "course_name": safe_text(course.name if course else "Unknown Course"),
            "teacher_name": safe_text(teacher.full_name if teacher else "Unassigned Teacher"),
            "room_name": safe_text(room.name if room else "Unassigned Room"),
            "group_name": safe_text(group.name if group else "Unassigned Group"),
            "course_type": "lab" if is_lab else "lecture",
            "status": "scheduled",
            "is_lab": is_lab,
            "course_color": course.color if course and course.color else "#0f8f5f",
        })
    return sort_master_rows(rows)

def detect_master_row_conflicts(rows):
    conflicts = []
    seen = {}
    for idx, row in enumerate(rows):
        for resource in ("teacher_name", "room_name", "group_name"):
            value = normalize_lookup_text(row[resource])
            if not value:
                continue
            key = (resource, value, row["day"], row["time_start"])
            if key in seen:
                conflicts.append((seen[key], idx, key))
            else:
                seen[key] = idx
    return conflicts

def build_timetable_context(schedule):
    entries = []
    for e in schedule:
        c = Course.query.get(e.course_id)
        t = User.query.get(e.teacher_id)
        r = Room.query.get(e.room_id)
        g = Group.query.get(e.group_id)
        entries.append({"entry": e, "course": c, "teacher": t, "room": r, "group": g})
    return entries

def detect_conflicts(schedule):
    conflicts = []
    seen = {}
    for e in schedule:
        keys = [
            f"teacher_{e.teacher_id}_{e.day}_{e.time_start}",
            f"room_{e.room_id}_{e.day}_{e.time_start}",
            f"group_{e.group_id}_{e.day}_{e.time_start}",
        ]
        for k in keys:
            if k in seen:
                conflicts.append((seen[k], e.id, k))
            else:
                seen[k] = e.id
    return conflicts

# ─────────────── CONTEXT PROCESSORS ───────────────
@app.context_processor
def inject_globals():
    university_name = session.get("university_name")
    if session.get("user_id"):
        user = User.query.get(session["user_id"])
        uni = University.query.get(user.university_id) if user and user.university_id else None
        if uni:
            university_name = uni.name
            session["university_name"] = university_name

    return {
        "current_year": datetime.now().year,
        "session_role": session.get("role"),
        "session_name": session.get("full_name"),
        "session_uni": university_name,
        "session_avatar": session.get("avatar"),
    }

# ─────────────── INIT ───────────────
with app.app_context():
    db.create_all()
    migrate_db()
    seed_database()
    # Assign groups to any existing students that don't have one yet
    try:
        student = User.query.filter_by(email="student@karabakh.edu.az").first()
        student2 = User.query.filter_by(email="leyla@karabakh.edu.az").first()
        kom24a = Group.query.filter_by(name="KOM24A").first()
        kom24b = Group.query.filter_by(name="KOM24B").first()
        if student and kom24a and not student.group_id:
            student.group_id = kom24a.id
        if student2 and kom24b and not student2.group_id:
            student2.group_id = kom24b.id
        db.session.commit()
    except Exception as ex:
        print(f"Group assignment skipped: {ex}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, host="0.0.0.0", port=port)
